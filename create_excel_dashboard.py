# -*- coding: utf-8 -*-
"""
YES24 베스트셀러 CSV 데이터를 전처리하고, Excel의 동적 수식, 스타일링, 차트를 
활용하여 고품질의 EDA 대시보드를 작성하는 파이썬 스크립트입니다.
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def create_excel_dashboard(csv_filepath, xlsx_filepath):
    print(f"데이터 로드 및 전처리 시작: {csv_filepath}")
    
    # 1. 데이터 로드 및 인코딩 처리
    df = pd.read_csv(csv_filepath, encoding='utf-8')
    
    # 2. 데이터 전처리
    # 포인트 컬럼 정제 (예: "1,200원" -> 1200)
    df['포인트'] = df['포인트'].astype(str).str.replace('원', '').str.replace(',', '').str.strip()
    df['포인트'] = pd.to_numeric(df['포인트'], errors='coerce').fillna(0).astype(int)
    
    # 할인율을 비율(소수)로 변환 (예: 10 -> 0.1)
    df['할인율'] = df['할인율(%)'] / 100.0
    
    # 결측치 처리
    df['저자'] = df['저자'].fillna('저자 미상')
    df['부제목'] = df['부제목'].fillna('')
    
    # 불필요한 원본 할인율 컬럼 제거 및 열 순서 재배치
    cols = ['순위', '상품ID', '상품명', '부제목', '저자', '출판사', '출판일', '할인율', '판매가', '정가', '포인트', '판매지수', '리뷰건수', '리뷰총점', '이미지URL']
    df = df[cols]
    
    # 출판사 Top 10 추출 (대시보드에 기입할 목적)
    top10_publishers = df['출판사'].value_counts().head(10).index.tolist()
    
    # 3. Excel Workbook 및 시트 생성
    wb = Workbook()
    
    # 기본 생성된 시트를 Dashboard로 변경
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    # EDA Summary 및 Raw Data 시트 추가
    ws_eda = wb.create_sheet(title="EDA_Summary")
    ws_raw = wb.create_sheet(title="Raw_Data")
    
    # 그리드선(눈금선) 표시 활성화
    for ws in [ws_dash, ws_eda, ws_raw]:
        ws.views.sheetView[0].showGridLines = True
        
    # --- 스타일 정의 ---
    font_family = "맑은 고딕"
    
    # 폰트 스타일
    font_title = Font(name=font_family, size=18, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=14, bold=True, color="1F4E79")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_body = Font(name=font_family, size=10, color="000000")
    font_body_bold = Font(name=font_family, size=10, bold=True, color="000000")
    
    # 채우기 스타일 (Navy 테마)
    fill_navy_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_light_navy = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    
    # 정렬 스타일
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 테두리 스타일
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all_thin = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
    
    double_bottom_side = Side(border_style="double", color="1F4E79")
    border_total = Border(top=thin_border_side, bottom=double_bottom_side)
    
    # 숫자 서식 정의 (규칙: 0은 "-"로 처리)
    fmt_currency = r"#,##0;(#,##0);'-'"
    fmt_percent = r"0.0%;(0.0%);'-'"
    fmt_integer = r"#,##0;(#,##0);'-'"
    fmt_decimal = r"0.0;(0.0);'-'"
    fmt_corr = r"0.00;(0.00);'-'"

    # ==========================================
    # 4. Raw_Data 시트 채우기 및 포맷 설정
    # ==========================================
    print("Raw_Data 시트 작성 중...")
    
    # 헤더 작성
    ws_raw.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_navy_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws_raw.row_dimensions[1].height = 25
    
    # 데이터 추가
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            cell.font = font_body
            cell.border = border_all_thin
            
            # 짝수행 얼룩무늬 효과
            if r_idx % 2 == 0:
                cell.fill = fill_zebra
                
            # 열별 데이터 서식 및 정렬 설정
            col_name = cols[c_idx - 1]
            if col_name in ['순위', '상품ID']:
                cell.alignment = align_center
            elif col_name in ['할인율']:
                cell.number_format = fmt_percent
                cell.alignment = align_right
            elif col_name in ['판매가', '정가', '포인트', '판매지수', '리뷰건수']:
                cell.number_format = fmt_currency
                cell.alignment = align_right
            elif col_name in ['리뷰총점']:
                cell.number_format = fmt_decimal
                cell.alignment = align_right
            else:
                cell.alignment = align_left
                
        ws_raw.row_dimensions[r_idx].height = 20

    # ==========================================
    # 5. EDA_Summary 시트 채우기 및 포맷 설정
    # ==========================================
    print("EDA_Summary 시트 작성 중...")
    
    # 5.1 기술통계 요약 테이블
    ws_eda.cell(row=2, column=2, value="수치형 데이터 기술통계 요약").font = font_section
    
    eda_headers = ["통계량", "할인율", "판매가 (원)", "정가 (원)", "포인트 (원)", "판매지수", "리뷰건수 (건)", "리뷰총점"]
    for col_idx, header in enumerate(eda_headers, 2):
        cell = ws_eda.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws_eda.row_dimensions[4].height = 24
    
    stats = [
        ("평균", "AVERAGE"),
        ("최솟값", "MIN"),
        ("중앙값", "MEDIAN"),
        ("최댓값", "MAX"),
        ("합계", "SUM"),
        ("데이터 수", "COUNT")
    ]
    
    # 수치형 변수들이 있는 Raw_Data의 열 번호 매핑
    raw_col_map = {
        "할인율": "H",
        "판매가 (원)": "I",
        "정가 (원)": "J",
        "포인트 (원)": "K",
        "판매지수": "L",
        "리뷰건수 (건)": "M",
        "리뷰총점": "N"
    }
    
    max_row = len(df) + 1 # 870
    
    for s_idx, (stat_name, excel_func) in enumerate(stats, 5):
        ws_eda.cell(row=s_idx, column=2, value=stat_name).font = font_body_bold
        ws_eda.cell(row=s_idx, column=2).alignment = align_left
        ws_eda.cell(row=s_idx, column=2).border = border_all_thin
        
        for c_idx, var_name in enumerate(eda_headers[1:], 3):
            raw_col = raw_col_map[var_name]
            formula = f"={excel_func}(Raw_Data!${raw_col}$2:${raw_col}${max_row})"
            
            cell = ws_eda.cell(row=s_idx, column=c_idx, value=formula)
            cell.font = font_body
            cell.border = border_all_thin
            
            # 서식 적용
            if var_name == "할인율":
                cell.number_format = fmt_percent
            elif var_name in ["판매가 (원)", "정가 (원)", "포인트 (원)", "판매지수", "리뷰건수 (건)"]:
                cell.number_format = fmt_currency
            elif var_name == "리뷰총점":
                cell.number_format = fmt_decimal
            cell.alignment = align_right
            
        ws_eda.row_dimensions[s_idx].height = 20
        
    # 기술통계 테이블 합계/평균 하단 더블 테두리 적용
    for col_idx in range(2, len(eda_headers) + 2):
        ws_eda.cell(row=10, column=col_idx).border = Border(
            top=thin_border_side, bottom=double_bottom_side,
            left=thin_border_side, right=thin_border_side
        )

    # 5.2 상관관계 매트릭스 테이블
    ws_eda.cell(row=13, column=2, value="변수 간 상관관계 매트릭스").font = font_section
    
    corr_vars = ["정가", "판매가", "할인율", "포인트", "판매지수", "리뷰건수", "리뷰총점"]
    corr_col_map = {
        "정가": "J",
        "판매가": "I",
        "할인율": "H",
        "포인트": "K",
        "판매지수": "L",
        "리뷰건수": "M",
        "리뷰총점": "N"
    }
    
    # 가로 헤더 작성
    ws_eda.cell(row=15, column=2, value="변수명").font = font_header
    ws_eda.cell(row=15, column=2).fill = fill_navy_header
    ws_eda.cell(row=15, column=2).alignment = align_center
    ws_eda.cell(row=15, column=2).border = border_all_thin
    
    for c_idx, var_name in enumerate(corr_vars, 3):
        cell = ws_eda.cell(row=15, column=c_idx, value=var_name)
        cell.font = font_header
        cell.fill = fill_navy_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws_eda.row_dimensions[15].height = 24
    
    # 세로 헤더 및 교차 수식 작성
    for r_idx, var_r in enumerate(corr_vars, 16):
        cell_r_header = ws_eda.cell(row=r_idx, column=2, value=var_r)
        cell_r_header.font = font_body_bold
        cell_r_header.alignment = align_left
        cell_r_header.border = border_all_thin
        
        col_r = corr_col_map[var_r]
        
        for c_idx, var_c in enumerate(corr_vars, 3):
            col_c = corr_col_map[var_c]
            formula = f"=CORREL(Raw_Data!${col_r}$2:${col_r}${max_row}, Raw_Data!${col_c}$2:${col_c}${max_row})"
            
            cell = ws_eda.cell(row=r_idx, column=c_idx, value=formula)
            cell.font = font_body
            cell.border = border_all_thin
            cell.number_format = fmt_corr
            cell.alignment = align_right
            
            # 대각선은 배경 연하게
            if var_r == var_c:
                cell.fill = fill_light_navy
                cell.font = font_body_bold
                
        ws_eda.row_dimensions[r_idx].height = 20

    # ==========================================
    # 6. Dashboard 시트 채우기 및 포맷 설정
    # ==========================================
    print("Dashboard 시트 작성 중...")
    
    # 6.1 타이틀
    ws_dash.merge_cells("B2:K2")
    title_cell = ws_dash.cell(row=2, column=2, value="YES24 IT/모바일 베스트셀러 데이터 분석 대시보드")
    title_cell.font = font_title
    title_cell.fill = fill_navy_header
    title_cell.alignment = align_center
    ws_dash.row_dimensions[2].height = 45
    
    # 6.2 KPI 카드 스타일링 헬퍼 함수 호출
    def make_kpi_card(start_col, title, formula, number_format):
        ws_dash.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col+1)
        ws_dash.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=start_col+1)
        
        # 타이틀
        t_cell = ws_dash.cell(row=4, column=start_col, value=title)
        t_cell.font = Font(name=font_family, size=9, bold=True, color="595959")
        t_cell.alignment = align_center
        
        # 값
        v_cell = ws_dash.cell(row=5, column=start_col, value=formula)
        v_cell.font = Font(name=font_family, size=16, bold=True, color="1F4E79")
        v_cell.alignment = align_center
        if number_format:
            v_cell.number_format = number_format
            
        # 테두리 및 배경색 처리
        title_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        val_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        
        for r in range(4, 7):
            for c in range(start_col, start_col + 2):
                cell = ws_dash.cell(row=r, column=c)
                cell.border = Border(
                    top=thin_border_side if r == 4 else None,
                    bottom=thin_border_side if r == 6 else None,
                    left=thin_border_side if c == start_col else None,
                    right=thin_border_side if c == start_col+1 else None
                )
                if r == 4:
                    cell.fill = title_fill
                else:
                    cell.fill = val_fill
                    
    # KPI 카드 추가
    make_kpi_card(2, "총 도서 수 (건)", f"=COUNTA(Raw_Data!$C$2:$C${max_row})", fmt_integer)
    make_kpi_card(4, "평균 정가 (원)", f"=AVERAGE(Raw_Data!$J$2:$J${max_row})", fmt_currency)
    make_kpi_card(6, "평균 할인율", f"=AVERAGE(Raw_Data!$H$2:$H${max_row})", fmt_percent)
    make_kpi_card(8, "총 판매지수", f"=SUM(Raw_Data!$L$2:$L${max_row})", fmt_integer)
    make_kpi_card(10, "평균 리뷰총점 (10점)", f"=AVERAGE(Raw_Data!$N$2:$N${max_row})", fmt_decimal)
    
    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 20
    ws_dash.row_dimensions[6].height = 20
    
    # 6.3 출판사별 요약 표 (Top 10)
    ws_dash.cell(row=8, column=2, value="출판사별 주요 지표 (Top 10)").font = font_section
    
    dash_tbl_headers = ["출판사명", "도서 수 (권)", "평균 판매지수"]
    for col_idx, header in enumerate(dash_tbl_headers, 2):
        cell = ws_dash.cell(row=9, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_navy_header
        cell.alignment = align_center
        cell.border = border_all_thin
    ws_dash.row_dimensions[9].height = 24
    
    for idx, pub in enumerate(top10_publishers, 10):
        # 출판사명 기입
        ws_dash.cell(row=idx, column=2, value=pub).font = font_body
        ws_dash.cell(row=idx, column=2).alignment = align_left
        ws_dash.cell(row=idx, column=2).border = border_all_thin
        
        # 도서 수 수식 기입
        formula_cnt = f'=COUNTIF(Raw_Data!$F$2:$F${max_row}, B{idx})'
        cell_cnt = ws_dash.cell(row=idx, column=3, value=formula_cnt)
        cell_cnt.font = font_body
        cell_cnt.number_format = fmt_integer
        cell_cnt.alignment = align_right
        cell_cnt.border = border_all_thin
        
        # 평균 판매지수 수식 기입
        formula_avg_sales = f'=AVERAGEIF(Raw_Data!$F$2:$F${max_row}, B{idx}, Raw_Data!$L$2:$L${max_row})'
        cell_avg = ws_dash.cell(row=idx, column=4, value=formula_avg_sales)
        cell_avg.font = font_body
        cell_avg.number_format = fmt_integer
        cell_avg.alignment = align_right
        cell_avg.border = border_all_thin
        
        ws_dash.row_dimensions[idx].height = 20
        
    # 요약 행 (Row 20)
    ws_dash.cell(row=20, column=2, value="합계 / 평균").font = font_body_bold
    ws_dash.cell(row=20, column=2).alignment = align_left
    ws_dash.cell(row=20, column=2).border = border_total
    
    cell_tot_cnt = ws_dash.cell(row=20, column=3, value="=SUM(C10:C19)")
    cell_tot_cnt.font = font_body_bold
    cell_tot_cnt.number_format = fmt_integer
    cell_tot_cnt.alignment = align_right
    cell_tot_cnt.border = border_total
    
    cell_avg_sales = ws_dash.cell(row=20, column=4, value="=AVERAGE(D10:D19)")
    cell_avg_sales.font = font_body_bold
    cell_avg_sales.number_format = fmt_integer
    cell_avg_sales.alignment = align_right
    cell_avg_sales.border = border_total
    ws_dash.row_dimensions[20].height = 22
    
    # 6.4 openpyxl 차트 생성 및 추가
    print("대시보드 차트 생성 중...")
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Top 10 출판사별 도서 수"
    chart.y_axis.title = "도서 수 (권)"
    chart.x_axis.title = "출판사"
    chart.width = 17
    chart.height = 10
    
    # 데이터(도서 수 C10:C19) 및 카테고리(출판사 B10:B19) 설정
    data_ref = Reference(ws_dash, min_col=3, min_row=9, max_row=19) # 헤더 포함하여 범위 지정
    cats_ref = Reference(ws_dash, min_col=2, min_row=10, max_row=19)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None # 범례 제거
    
    # 차트 삽입 위치
    ws_dash.add_chart(chart, "F8")

    # ==========================================
    # 7. 열 너비 자동 조정
    # ==========================================
    print("열 너비 자동 조정 중...")
    for ws in [ws_dash, ws_eda, ws_raw]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "1,234,567"
                
                # 한글 글자 길이 감안하여 바이트 수 계산
                val_len = 0
                for char in val_str:
                    if ord(char) > 127: # 한글 등 멀티바이트
                        val_len += 2
                    else:
                        val_len += 1
                
                if val_len > max_len:
                    max_len = val_len
            
            # 마진 추가 및 너비 설정 (최소 10, 최대 45로 제한)
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 45), 10)
            
    # Dashboard 특정 열 너비 미세 조정
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 24
    ws_dash.column_dimensions['C'].width = 15
    ws_dash.column_dimensions['D'].width = 18
    
    # 8. 저장
    print(f"Excel 파일 저장 중: {xlsx_filepath}")
    wb.save(xlsx_filepath)
    print("Excel 대시보드 생성 완료!")

if __name__ == "__main__":
    csv_path = os.path.join("yes24", "data", "yes24_bestseller.csv")
    xlsx_path = os.path.join("yes24", "data", "yes24_bestseller_eda.xlsx")
    
    if os.path.exists(csv_path):
        create_excel_dashboard(csv_path, xlsx_path)
    else:
        print(f"오류: CSV 파일이 존재하지 않습니다: {csv_path}")
